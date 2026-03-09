"""
Module de parsing des grilles tarifaires Excel MINT Énergie 2026
Parse les 3 onglets : Grille_C2, Grille_C4, Grille_C5
"""
import openpyxl
from datetime import datetime
from pathlib import Path


class ExcelGrilleParser:
    """Parser pour les grilles tarifaires Excel MINT"""

    def __init__(self, excel_path):
        """
        Initialise le parser avec le fichier Excel

        Args:
            excel_path: Chemin vers le fichier .xlsx
        """
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.grilles = {
            'C2': None,
            'C4': None,
            'C5': None
        }
        self.metadata = {}
        self._load_excel()

    def _load_excel(self):
        """Charge le fichier Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            print(f"✅ Excel chargé : {self.excel_path.name}")
            print(f"   Onglets disponibles : {self.workbook.sheetnames}")
        except Exception as e:
            print(f"❌ Erreur chargement Excel : {e}")
            raise

    def parse_all(self):
        """Parse les 3 onglets de grilles"""
        for segment in ['C2', 'C4', 'C5']:
            sheet_name = f'Grille_{segment}'
            if sheet_name in self.workbook.sheetnames:
                self.grilles[segment] = self._parse_onglet(segment, sheet_name)
                print(f"✅ Grille {segment} parsée : {len(self.grilles[segment])} lignes")
            else:
                print(f"⚠️ Onglet {sheet_name} non trouvé")

        return self.grilles

    def _parse_onglet(self, segment, sheet_name):
        """
        Parse un onglet de grille

        Args:
            segment: 'C2', 'C4' ou 'C5'
            sheet_name: Nom de l'onglet Excel

        Returns:
            list: Liste des lignes de grille
        """
        sheet = self.workbook[sheet_name]
        lignes = []

        # Extraire métadonnées (lignes 2-4)
        date_grille = sheet['B2'].value
        seuil = sheet['B4'].value

        self.metadata[segment] = {
            'date_grille': self._parse_date_cell(date_grille),
            'seuil': str(seuil) if seuil else ''
        }

        # Ligne 13 = En-têtes (pour validation)
        # Ligne 14 = Unités
        # Lignes 15+ = Données

        row_num = 15  # Commence à la ligne 15

        while True:
            # Lire la ligne
            date_debut_cell = sheet[f'A{row_num}'].value

            # Si cellule vide, fin des données
            if not date_debut_cell:
                break

            try:
                ligne = self._parse_ligne(sheet, row_num, segment)
                if ligne:
                    lignes.append(ligne)
            except Exception as e:
                print(f"⚠️ Erreur ligne {row_num} ({segment}): {e}")

            row_num += 1

            # Limite de sécurité
            if row_num > 1000:
                break

        return lignes

    def _parse_ligne(self, sheet, row_num, segment):
        """
        Parse une ligne de données selon le segment

        Args:
            sheet: Feuille Excel
            row_num: Numéro de ligne
            segment: 'C2', 'C4' ou 'C5'

        Returns:
            dict: Données de la ligne
        """
        # Colonnes communes A-C
        date_debut = self._parse_date_cell(sheet[f'A{row_num}'].value)
        duree_mois = sheet[f'B{row_num}'].value
        date_fin = self._parse_date_cell(sheet[f'C{row_num}'].value)

        if not date_debut or not date_fin:
            return None

        ligne = {
            'date_debut': date_debut,
            'duree_mois': int(duree_mois) if duree_mois else 0,
            'date_fin': date_fin,
        }

        # Prix selon le segment
        if segment == 'C2':
            # Colonnes D-I : PTE, HPH, HCH, HPE, HCE, Coefficient α
            ligne['prix_pte'] = self._parse_float(sheet[f'D{row_num}'].value)
            ligne['prix_hph'] = self._parse_float(sheet[f'E{row_num}'].value)
            ligne['prix_hch'] = self._parse_float(sheet[f'F{row_num}'].value)
            ligne['prix_hpe'] = self._parse_float(sheet[f'G{row_num}'].value)
            ligne['prix_hce'] = self._parse_float(sheet[f'H{row_num}'].value)
            ligne['coefficient_alpha'] = self._parse_float(sheet[f'I{row_num}'].value)

        elif segment == 'C4':
            # Colonnes D-H : HPH, HCH, HPE, HCE, Coefficient α
            ligne['prix_hph'] = self._parse_float(sheet[f'D{row_num}'].value)
            ligne['prix_hch'] = self._parse_float(sheet[f'E{row_num}'].value)
            ligne['prix_hpe'] = self._parse_float(sheet[f'F{row_num}'].value)
            ligne['prix_hce'] = self._parse_float(sheet[f'G{row_num}'].value)
            ligne['coefficient_alpha'] = self._parse_float(sheet[f'H{row_num}'].value)

        elif segment == 'C5':
            # Colonnes D-K : Base, HP, HC, HPSH, HCSH, HPSB, HCSB, Coefficient α
            ligne['prix_base'] = self._parse_float(sheet[f'D{row_num}'].value)
            ligne['prix_hp'] = self._parse_float(sheet[f'E{row_num}'].value)
            ligne['prix_hc'] = self._parse_float(sheet[f'F{row_num}'].value)
            ligne['prix_hpsh'] = self._parse_float(sheet[f'G{row_num}'].value)
            ligne['prix_hcsh'] = self._parse_float(sheet[f'H{row_num}'].value)
            ligne['prix_hpsb'] = self._parse_float(sheet[f'I{row_num}'].value)
            ligne['prix_hcsb'] = self._parse_float(sheet[f'J{row_num}'].value)
            ligne['coefficient_alpha'] = self._parse_float(sheet[f'K{row_num}'].value)

        return ligne

    def _parse_date_cell(self, cell_value):
        """
        Parse une cellule de date (peut être datetime ou string)

        Returns:
            str: Date au format DD/MM/YYYY
        """
        if not cell_value:
            return None

        if isinstance(cell_value, datetime):
            return cell_value.strftime('%d/%m/%Y')

        if isinstance(cell_value, str):
            # Déjà au bon format
            if '/' in cell_value:
                return cell_value
            # Format ISO
            try:
                dt = datetime.strptime(cell_value, '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y')
            except:
                pass

        return str(cell_value)

    def _parse_float(self, value):
        """Parse une valeur float avec gestion des erreurs"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except:
            return 0.0

    def get_dates_disponibles(self, segment):
        """
        Récupère toutes les dates de début disponibles pour un segment

        Args:
            segment: 'C2', 'C4' ou 'C5'

        Returns:
            list: Liste des dates de début uniques (format DD/MM/YYYY), triées chronologiquement
        """
        if not self.grilles[segment]:
            return []

        dates = set()
        for ligne in self.grilles[segment]:
            dates.add(ligne['date_debut'])

        # FIX BUG 1: Trier chronologiquement (datetime) au lieu d'alphabétiquement (string)
        # Convertir en datetime pour tri, puis reconvertir en string DD/MM/YYYY
        from datetime import datetime
        dates_list = list(dates)
        dates_sorted = sorted(dates_list, key=lambda d: datetime.strptime(d, '%d/%m/%Y'))
        return dates_sorted

    def get_durees_disponibles(self, segment, date_debut):
        """
        Récupère les durées disponibles pour un segment et une date de début

        Args:
            segment: 'C2', 'C4' ou 'C5'
            date_debut: Date de début (format DD/MM/YYYY)

        Returns:
            list: Liste des durées disponibles (en mois)
        """
        if not self.grilles[segment]:
            return []

        durees = []
        for ligne in self.grilles[segment]:
            if ligne['date_debut'] == date_debut:
                durees.append(ligne['duree_mois'])

        return sorted(durees)

    def get_prix_p0(self, segment, date_debut, duree_mois):
        """
        Récupère les prix P0 pour un segment, date et durée donnés
        Si la date exacte n'existe pas, cherche par mois/année
        Si la durée exacte n'existe pas, prend la durée la plus proche disponible

        Args:
            segment: 'C2', 'C4' ou 'C5'
            date_debut: Date de début (format DD/MM/YYYY)
            duree_mois: Durée en mois

        Returns:
            dict: Ligne de grille avec tous les prix, ou None
        """
        if not self.grilles[segment]:
            return None

        from datetime import datetime

        # DEBUG BUG 2: Afficher les dates de la grille vs date recherchée
        print(f"🔍 Recherche prix: segment={segment}, date_debut='{date_debut}' (type={type(date_debut)}), duree_mois={duree_mois}")

        # 1. Chercher d'abord un match exact (date + durée)
        for ligne in self.grilles[segment]:
            if (ligne['date_debut'] == date_debut and
                ligne['duree_mois'] == duree_mois):
                print(f"✅ Match exact trouvé: {date_debut} - {duree_mois} mois")
                return ligne

        # 2. Si pas de match exact, chercher par date et prendre la durée la plus proche
        print(f"⚠️ Pas de match exact, recherche de la durée la plus proche...")
        lignes_meme_date = [ligne for ligne in self.grilles[segment] if ligne['date_debut'] == date_debut]

        if lignes_meme_date:
            # Prendre la ligne avec la durée la plus proche (priorité : plus grande si égale distance)
            ligne_proche = min(lignes_meme_date, key=lambda r: abs(r['duree_mois'] - duree_mois))
            duree_utilisee = ligne_proche['duree_mois']
            print(f"✅ Durée {duree_mois} mois non disponible, utilisation de {duree_utilisee} mois (durée la plus proche)")
            return ligne_proche

        # 3. FIX BUG 2: Si la date exacte n'existe pas, chercher par MOIS/ANNÉE
        # Ex: si on cherche 01/03/2026 mais la grille a 04/03/2026, utiliser 04/03/2026
        print(f"⚠️ Date {date_debut} non trouvée, recherche par mois/année...")
        try:
            date_recherchee = datetime.strptime(date_debut, '%d/%m/%Y')
            target_month = date_recherchee.month
            target_year = date_recherchee.year

            # Chercher toutes les lignes du même mois/année
            lignes_meme_mois = [
                ligne for ligne in self.grilles[segment]
                if datetime.strptime(ligne['date_debut'], '%d/%m/%Y').month == target_month
                and datetime.strptime(ligne['date_debut'], '%d/%m/%Y').year == target_year
            ]

            if lignes_meme_mois:
                # Prendre la ligne avec la durée la plus proche
                ligne_proche = min(lignes_meme_mois, key=lambda r: abs(r['duree_mois'] - duree_mois))
                date_utilisee = ligne_proche['date_debut']
                duree_utilisee = ligne_proche['duree_mois']
                print(f"✅ Date {date_debut} non trouvée, utilisation de {date_utilisee} (même mois) avec durée {duree_utilisee} mois")
                return ligne_proche
        except Exception as e:
            print(f"⚠️ Erreur parsing date: {e}")

        # 4. Si aucun match, retourner None
        print(f"❌ Aucun prix trouvé pour {date_debut} (ni date exacte, ni même mois)")
        return None

    def calculer_prix_avec_marge(self, prix_p0, marge_courtier):
        """
        Calcule les prix finaux avec marge courtier

        Args:
            prix_p0: dict des prix P0 (retour de get_prix_p0)
            marge_courtier: Marge courtier en €/MWh (min 6, max 25)

        Returns:
            dict: Prix finaux arrondis à 2 décimales
        """
        # Limiter la marge
        marge_courtier = max(6.0, min(25.0, float(marge_courtier)))

        prix_finaux = {}

        for key, value in prix_p0.items():
            if key.startswith('prix_') and key != 'prix_capa_2025' and key != 'prix_capa_2026':
                try:
                    p0 = float(value)
                    prix_final = p0 + marge_courtier
                    prix_finaux[key] = round(prix_final, 2)
                except:
                    prix_finaux[key] = value
            else:
                # Conserver les autres champs (dates, durée, coefficient)
                prix_finaux[key] = value

        # Ajouter les infos de marge
        prix_finaux['marge_courtier'] = round(marge_courtier, 2)

        return prix_finaux

    def get_metadata(self, segment=None):
        """
        Récupère les métadonnées d'un segment ou de tous

        Args:
            segment: 'C2', 'C4', 'C5' ou None (tous)

        Returns:
            dict: Métadonnées
        """
        if segment:
            return self.metadata.get(segment, {})
        return self.metadata


if __name__ == '__main__':
    # Test avec un fichier Excel
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <chemin_fichier.xlsx>")
        sys.exit(1)

    excel_path = sys.argv[1]

    try:
        parser = ExcelGrilleParser(excel_path)
        grilles = parser.parse_all()

        print("\n" + "="*60)
        print("MÉTADONNÉES")
        print("="*60)
        for segment, meta in parser.metadata.items():
            print(f"\n{segment}:")
            print(f"  Date grille : {meta.get('date_grille')}")
            print(f"  Seuil : {meta.get('seuil')}")

        print("\n" + "="*60)
        print("TEST RÉCUPÉRATION PRIX")
        print("="*60)

        # Test C4
        dates = parser.get_dates_disponibles('C4')
        if dates:
            date_test = dates[0]
            durees = parser.get_durees_disponibles('C4', date_test)
            if durees:
                duree_test = durees[0]
                prix = parser.get_prix_p0('C4', date_test, duree_test)

                print(f"\n✅ Test C4 : {date_test} - {duree_test} mois")
                print(f"   Prix P0 HPH : {prix['prix_hph']:.2f} €/MWh")
                print(f"   Prix P0 HCH : {prix['prix_hch']:.2f} €/MWh")
                print(f"   Coefficient α : {prix['coefficient_alpha']}")

                # Test avec marge
                prix_marge = parser.calculer_prix_avec_marge(prix, 10.0)
                print(f"\n💰 Avec marge 10 €/MWh :")
                print(f"   Prix final HPH : {prix_marge['prix_hph']:.2f} €/MWh")
                print(f"   Prix final HCH : {prix_marge['prix_hch']:.2f} €/MWh")

    except Exception as e:
        print(f"❌ Erreur : {e}")
        import traceback
        traceback.print_exc()
